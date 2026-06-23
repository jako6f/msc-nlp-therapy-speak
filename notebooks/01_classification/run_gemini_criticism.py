#!/usr/bin/env python3
"""Run reproducible, schema-validated Gemini frame criticism."""

from __future__ import annotations

import argparse
import hashlib
import json
import numbers
import re
import shutil
import subprocess
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import run_codex_annotation as codex_runner
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.metrics import average_precision_score

PROJECT_ROOT = Path(__file__).resolve().parents[2]
CLASSIFICATION_DIR = PROJECT_ROOT / "data/interim/lsc/classification"
LLM_DIR = CLASSIFICATION_DIR / "llm_annotation"
CODEX_DIR = LLM_DIR / "codex"
GEMINI_DIR = LLM_DIR / "gemini"
CODEX_PILOT_DIR = CODEX_DIR / "pilot"
CODEX_PRODUCTION_DIR = CODEX_DIR / "production"
HUMAN_DIR = CLASSIFICATION_DIR / "human_annotation"
CORRECTION_DIR = CLASSIFICATION_DIR / "human_correction"
PROMPT_DIR = PROJECT_ROOT / "notebooks/01_classification/prompts"

CODEBOOK_PATH = PROJECT_ROOT / "notebooks/01_classification/codebooks/codebook_v4.md"
PROMPT_PATH = PROMPT_DIR / "critic_v5.md"
SCHEMA_PATH = PROMPT_DIR / "critic_output_schema.json"
NO_TOOLS_POLICY_PATH = PROMPT_DIR / "gemini_critic_no_tools.toml"
PROMPT_VERSION = "critic_v5"
CODEBOOK_VERSION = "v0.4"
ANNOTATOR_REASONING = "high"
CRITIC_BATCH_SIZE = 10
DEFAULT_MODEL = "gemini-3-flash-preview"

GEMINI_PILOT_EVAL = GEMINI_DIR / "pilot" / "evaluation"

ANNOTATOR_BATCH_DIR = CODEX_PRODUCTION_DIR / "inputs"
ANNOTATOR_RAW_DIR = CODEX_PRODUCTION_DIR / "outputs"
ANNOTATOR_MANIFEST = CODEX_PRODUCTION_DIR / "manifest.csv"
ANNOTATOR_POOL = CODEX_PRODUCTION_DIR / "training_pool_with_annotation_ids.csv"
ANNOTATOR_LABELS = CODEX_PRODUCTION_DIR / "labels.csv"
PILOT_BATCH_DIR = CODEX_PILOT_DIR / "inputs"
PILOT_ANNOTATOR_RAW_DIR = CODEX_PILOT_DIR / "outputs" / ANNOTATOR_REASONING
PILOT_MANIFEST = CODEX_PILOT_DIR / "manifest.csv"
PILOT_SOURCE = HUMAN_DIR / "frame_pilot_annotation.xlsx"

OUTPUT_FIELDS = [
    "annotation_id",
    "overall_error_prob",
    "substantive_error_prob",
    "clinical_error_prob",
    "lived_error_prob",
    "substantive_suggested_label",
    "clinical_suggested_label",
    "lived_suggested_label",
    "reason",
    "evidence",
]
INPUT_FIELDS = [
    "annotation_id",
    "context_id",
    "target",
    "raw_form",
    "passage",
    "proposed_substantive_target_discourse",
    "proposed_clinical_frame_present",
    "proposed_lived_experience_frame_present",
    "proposed_confidence",
]
CONFIDENCE_VALUES = {"high", "medium", "low"}
REVIEW_STATUS_VALUES = {"pending", "reviewed"}
LABEL_COLUMNS = [
    "substantive_target_discourse",
    "clinical_frame_present",
    "lived_experience_frame_present",
]


class NonRetryableGeminiError(RuntimeError):
    """Gemini error that should stop the current run immediately."""


def utc_now() -> str:
    return datetime.now(UTC).isoformat()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def model_slug(model: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", model).strip("_")


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows = []
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as error:
                raise ValueError(
                    f"Invalid JSON in {path.name} line {line_number}: {error}"
                ) from error
            if not isinstance(row, dict):
                raise ValueError(f"{path.name} line {line_number}: expected an object.")
            rows.append(row)
    return rows


def write_jsonl(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")


def parse_bool_or_none(value: Any) -> bool | None:
    if value is None or pd.isna(value):
        return None
    if type(value) is bool:
        return value
    if isinstance(value, numbers.Number):
        if value == 1:
            return True
        if value == 0:
            return False
    text = str(value).strip().lower()
    if text in {"", "na", "n/a", "none", "null", "nan"}:
        return None
    if text in {"true", "t", "yes", "y", "1"}:
        return True
    if text in {"false", "f", "no", "n", "0"}:
        return False
    raise ValueError(f"Cannot parse boolean value: {value!r}")


def label_text(value: bool | None) -> str:
    if value is None:
        return "NA"
    return "TRUE" if value else "FALSE"


def validate_hierarchy(substantive: bool, clinical: bool | None, lived: bool | None) -> None:
    if substantive and (clinical is None or lived is None):
        raise ValueError("Substantive rows require boolean clinical and lived labels.")
    if not substantive and (clinical is not None or lived is not None):
        raise ValueError("Non-substantive rows require null clinical and lived labels.")


def derive_frame(substantive: bool, clinical: bool | None, lived: bool | None) -> str:
    validate_hierarchy(substantive, clinical, lived)
    if not substantive:
        return "non_substantive_or_insufficient"
    if clinical and lived:
        return "mixed"
    if clinical:
        return "clinical_only"
    if lived:
        return "lived_only"
    return "substantive_other"


def dataset_input_dir(dataset: str) -> Path:
    return GEMINI_DIR / dataset / "inputs"


def dataset_manifest_path(dataset: str) -> Path:
    return GEMINI_DIR / dataset / "manifest.csv"


def source_batch_dir(dataset: str) -> Path:
    return PILOT_BATCH_DIR if dataset == "pilot" else ANNOTATOR_BATCH_DIR


def source_raw_dir(dataset: str) -> Path:
    return PILOT_ANNOTATOR_RAW_DIR if dataset == "pilot" else ANNOTATOR_RAW_DIR


def source_manifest_path(dataset: str) -> Path:
    return PILOT_MANIFEST if dataset == "pilot" else ANNOTATOR_MANIFEST


def source_batch_number(source_name: str) -> int:
    match = re.fullmatch(r"(?:annotator|pilot)_batch_(\d+)\.jsonl", source_name)
    if not match:
        raise ValueError(f"Unexpected source batch name: {source_name}")
    return int(match.group(1))


def critic_batch_name(source_name: str, chunk_number: int) -> str:
    return f"critic_batch_{source_batch_number(source_name):03d}_{chunk_number:02d}.jsonl"


def batch_selector_pattern(value: str) -> re.Pattern[str]:
    value = value.removesuffix(".jsonl")
    critic_match = re.fullmatch(r"critic_batch_(\d+)_(\d+)", value)
    if critic_match:
        return re.compile(
            rf"critic_batch_{int(critic_match.group(1)):03d}_{int(critic_match.group(2)):02d}\.jsonl"
        )
    source_match = re.fullmatch(r"(?:(?:critic|annotator|pilot)_batch_)?(\d+)", value)
    if not source_match:
        raise ValueError(
            "Batch must be a source-batch number/name or an exact "
            "critic_batch_<source>_<chunk> name."
        )
    return re.compile(rf"critic_batch_{int(source_match.group(1)):03d}_\d{{2}}\.jsonl")


def validate_source_input(path: Path) -> list[dict[str, Any]]:
    rows = read_jsonl(path)
    required = {"annotation_id", "context_id", "target", "raw_form", "passage"}
    ids = []
    for index, row in enumerate(rows, start=1):
        if set(row) != required:
            raise ValueError(
                f"{path.name} row {index}: required fields are exactly {sorted(required)}."
            )
        if not all(isinstance(row[key], str) for key in required):
            raise ValueError(f"{path.name} row {index}: every field must be a string.")
        ids.append(row["annotation_id"])
    if len(ids) != len(set(ids)):
        raise ValueError(f"{path.name}: duplicate annotation IDs.")
    return rows


def validate_annotator_output(
    rows: list[dict[str, Any]], expected_rows: list[dict[str, Any]], source_name: str
) -> list[dict[str, Any]]:
    fields = {
        "annotation_id",
        "substantive_target_discourse",
        "clinical_frame_present",
        "lived_experience_frame_present",
        "confidence",
    }
    if len(rows) != len(expected_rows):
        raise ValueError(f"{source_name}: expected {len(expected_rows)} rows, found {len(rows)}.")
    for index, row in enumerate(rows, start=1):
        if set(row) != fields:
            raise ValueError(f"{source_name} row {index}: invalid annotator output fields.")
        substantive = row["substantive_target_discourse"]
        clinical = row["clinical_frame_present"]
        lived = row["lived_experience_frame_present"]
        if type(substantive) is not bool:
            raise ValueError(f"{source_name} row {index}: substantive label must be boolean.")
        if clinical is not None and type(clinical) is not bool:
            raise ValueError(f"{source_name} row {index}: clinical label must be boolean or null.")
        if lived is not None and type(lived) is not bool:
            raise ValueError(f"{source_name} row {index}: lived label must be boolean or null.")
        validate_hierarchy(substantive, clinical, lived)
        if row["confidence"] not in CONFIDENCE_VALUES:
            raise ValueError(f"{source_name} row {index}: invalid confidence.")
    expected_ids = [row["annotation_id"] for row in expected_rows]
    actual_ids = [row["annotation_id"] for row in rows]
    if actual_ids != expected_ids:
        raise ValueError(f"{source_name}: annotation IDs or order differ from the source batch.")
    return rows


def critic_input_rows(
    source_rows: list[dict[str, Any]], annotation_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    return [
        {
            "annotation_id": source["annotation_id"],
            "context_id": source["context_id"],
            "target": source["target"],
            "raw_form": source["raw_form"],
            "passage": source["passage"],
            "proposed_substantive_target_discourse": annotation["substantive_target_discourse"],
            "proposed_clinical_frame_present": annotation["clinical_frame_present"],
            "proposed_lived_experience_frame_present": annotation["lived_experience_frame_present"],
            "proposed_confidence": annotation["confidence"],
        }
        for source, annotation in zip(source_rows, annotation_rows)
    ]


def validate_critic_input(path: Path) -> list[dict[str, Any]]:
    rows = read_jsonl(path)
    ids = []
    for index, row in enumerate(rows, start=1):
        if list(row) != INPUT_FIELDS:
            raise ValueError(
                f"{path.name} row {index}: required fields and order are exactly {INPUT_FIELDS}."
            )
        substantive = row["proposed_substantive_target_discourse"]
        clinical = row["proposed_clinical_frame_present"]
        lived = row["proposed_lived_experience_frame_present"]
        if type(substantive) is not bool:
            raise ValueError(f"{path.name} row {index}: proposed substantive must be boolean.")
        if clinical is not None and type(clinical) is not bool:
            raise ValueError(f"{path.name} row {index}: proposed clinical must be boolean or null.")
        if lived is not None and type(lived) is not bool:
            raise ValueError(f"{path.name} row {index}: proposed lived must be boolean or null.")
        validate_hierarchy(substantive, clinical, lived)
        if row["proposed_confidence"] not in CONFIDENCE_VALUES:
            raise ValueError(f"{path.name} row {index}: invalid proposed confidence.")
        ids.append(row["annotation_id"])
    if len(ids) != len(set(ids)):
        raise ValueError(f"{path.name}: duplicate annotation IDs.")
    return rows


def validate_critic_rows(
    criticisms: Any, expected_rows: list[dict[str, Any]], source_name: str
) -> list[dict[str, Any]]:
    if not isinstance(criticisms, list):
        raise ValueError(f"{source_name}: criticisms must be an array.")
    if len(criticisms) != len(expected_rows):
        raise ValueError(
            f"{source_name}: expected {len(expected_rows)} rows, found {len(criticisms)}."
        )
    for index, (row, proposed) in enumerate(zip(criticisms, expected_rows), start=1):
        if not isinstance(row, dict) or set(row) != set(OUTPUT_FIELDS):
            raise ValueError(f"{source_name} row {index}: required fields are {OUTPUT_FIELDS}.")
        for field in ["overall_error_prob", "substantive_error_prob"]:
            if type(row[field]) not in {int, float} or not 0 <= row[field] <= 1:
                raise ValueError(f"{source_name} row {index}: invalid {field}.")
        for field in ["clinical_error_prob", "lived_error_prob"]:
            value = row[field]
            if value is not None and (type(value) not in {int, float} or not 0 <= value <= 1):
                raise ValueError(f"{source_name} row {index}: invalid {field}.")
        proposed_substantive = proposed["proposed_substantive_target_discourse"]
        stage1_probs = [row["clinical_error_prob"], row["lived_error_prob"]]
        if proposed_substantive and any(value is None for value in stage1_probs):
            raise ValueError(
                f"{source_name} row {index}: substantive proposal needs Stage-1 scores."
            )
        if not proposed_substantive and any(value is not None for value in stage1_probs):
            raise ValueError(
                f"{source_name} row {index}: non-substantive proposal needs null Stage-1 scores."
            )
        suggested = [
            row["substantive_suggested_label"],
            row["clinical_suggested_label"],
            row["lived_suggested_label"],
        ]
        if type(suggested[0]) is not bool:
            raise ValueError(f"{source_name} row {index}: suggested substantive must be boolean.")
        if suggested[1] is not None and type(suggested[1]) is not bool:
            raise ValueError(f"{source_name} row {index}: suggested clinical is invalid.")
        if suggested[2] is not None and type(suggested[2]) is not bool:
            raise ValueError(f"{source_name} row {index}: suggested lived is invalid.")
        validate_hierarchy(*suggested)
        if not isinstance(row["reason"], str) or not isinstance(row["evidence"], str):
            raise ValueError(f"{source_name} row {index}: reason and evidence must be strings.")
        if len(row["reason"]) > 240 or len(row["evidence"]) > 180:
            raise ValueError(f"{source_name} row {index}: reason or evidence exceeds its limit.")
    expected_ids = [row["annotation_id"] for row in expected_rows]
    actual_ids = [row["annotation_id"] for row in criticisms]
    if actual_ids != expected_ids:
        raise ValueError(f"{source_name}: annotation IDs or order differ from critic input.")
    return criticisms


def normalise_critic_text(criticisms: Any) -> tuple[Any, int]:
    if not isinstance(criticisms, list):
        return criticisms, 0
    normalised = []
    changes = 0
    for row in criticisms:
        if not isinstance(row, dict):
            normalised.append(row)
            continue
        row = row.copy()
        for field, limit in {"reason": 240, "evidence": 180}.items():
            value = row.get(field)
            if isinstance(value, str) and len(value) > limit:
                row[field] = value[:limit].rstrip()
                changes += 1
        normalised.append(row)
    return normalised, changes


def normalise_hierarchy_nulls(
    criticisms: Any, expected_rows: list[dict[str, Any]]
) -> tuple[Any, int]:
    if not isinstance(criticisms, list) or len(criticisms) != len(expected_rows):
        return criticisms, 0
    normalised = []
    changes = 0
    for row, proposed in zip(criticisms, expected_rows):
        if not isinstance(row, dict):
            normalised.append(row)
            continue
        row = row.copy()
        if not proposed["proposed_substantive_target_discourse"]:
            for field in ["clinical_error_prob", "lived_error_prob"]:
                if row.get(field) is not None:
                    row[field] = None
                    changes += 1
        if row.get("substantive_suggested_label") is False:
            for field in ["clinical_suggested_label", "lived_suggested_label"]:
                if row.get(field) is not None:
                    row[field] = None
                    changes += 1
        normalised.append(row)
    return normalised, changes


def sync_inputs(dataset: str, dry_run: bool = False) -> list[Path]:
    source_manifest_file = source_manifest_path(dataset)
    if not source_manifest_file.exists():
        raise FileNotFoundError(f"Missing source manifest: {source_manifest_file}")
    source_manifest = pd.read_csv(source_manifest_file)
    input_dir = dataset_input_dir(dataset)
    input_dir.mkdir(parents=True, exist_ok=True)

    existing = (
        pd.read_csv(dataset_manifest_path(dataset))
        if dataset_manifest_path(dataset).exists()
        else pd.DataFrame()
    )
    existing_by_critic = (
        existing.set_index("critic_batch_name").to_dict("index") if not existing.empty else {}
    )
    manifest_rows = []
    paths = []
    for source_record in source_manifest.itertuples(index=False):
        source_name = str(source_record.batch_name)
        source_input = source_batch_dir(dataset) / source_name
        source_output = source_raw_dir(dataset) / source_name
        if not source_input.exists() or not source_output.exists():
            raise FileNotFoundError(
                f"Every declared source batch must be complete before criticism: {source_name}"
            )
        if hasattr(source_record, "sha256") and sha256(source_input) != str(source_record.sha256):
            raise ValueError(f"Source input differs from annotation manifest: {source_name}")
        if not codex_runner.validated_canonical_exists(dataset, ANNOTATOR_REASONING, source_input):
            raise ValueError(
                "Codex output is not current and validated under its recorded contract: "
                f"{source_name}"
            )
        source_rows = validate_source_input(source_input)
        annotation_rows = validate_annotator_output(
            read_jsonl(source_output), source_rows, source_output.name
        )
        complete_rows = critic_input_rows(source_rows, annotation_rows)
        source_input_hash = sha256(source_input)
        source_output_hash = sha256(source_output)
        for chunk_number, start in enumerate(
            range(0, len(complete_rows), CRITIC_BATCH_SIZE), start=1
        ):
            expected_rows = complete_rows[start : start + CRITIC_BATCH_SIZE]
            critic_name = critic_batch_name(source_name, chunk_number)
            critic_path = input_dir / critic_name
            old = existing_by_critic.get(critic_name)
            if old:
                expected_contract = {
                    "source_batch_name": source_name,
                    "source_chunk_number": chunk_number,
                    "critic_batch_size": CRITIC_BATCH_SIZE,
                    "rows": len(expected_rows),
                    "source_input_sha256": source_input_hash,
                    "source_output_sha256": source_output_hash,
                    "prompt_version": PROMPT_VERSION,
                    "codebook_version": CODEBOOK_VERSION,
                }
                differences = [
                    key
                    for key, value in expected_contract.items()
                    if str(old.get(key)) != str(value)
                ]
                if differences or not critic_path.exists():
                    raise ValueError(
                        f"Stale critic input for {critic_name}; run rebuild-batch explicitly. "
                        f"Differences: {differences or ['missing critic input']}"
                    )
                current_rows = validate_critic_input(critic_path)
                if sha256(critic_path) != str(old.get("critic_input_sha256")):
                    raise ValueError(
                        f"Critic input hash differs from its manifest for {critic_name}; "
                        "run rebuild-batch explicitly."
                    )
                if current_rows != expected_rows:
                    raise ValueError(
                        f"Critic input differs from current Codex labels for {critic_name}; "
                        "run rebuild-batch explicitly."
                    )
            else:
                if critic_path.exists():
                    raise ValueError(f"Unmanifested critic input already exists: {critic_path}")
                if not dry_run:
                    write_jsonl(expected_rows, critic_path)

            manifest_rows.append(
                {
                    "critic_batch_name": critic_name,
                    "source_batch_name": source_name,
                    "source_chunk_number": chunk_number,
                    "critic_batch_size": CRITIC_BATCH_SIZE,
                    "rows": len(expected_rows),
                    "annotation_id_start": expected_rows[0]["annotation_id"],
                    "annotation_id_end": expected_rows[-1]["annotation_id"],
                    "source_input_sha256": source_input_hash,
                    "source_output_sha256": source_output_hash,
                    "critic_input_sha256": sha256(critic_path)
                    if critic_path.exists()
                    else "pending",
                    "prompt_version": PROMPT_VERSION,
                    "codebook_version": CODEBOOK_VERSION,
                }
            )
            paths.append(critic_path)

    expected_names = {row["critic_batch_name"] for row in manifest_rows}
    unexpected_names = set(existing_by_critic) - expected_names
    if unexpected_names:
        raise ValueError(
            "Critic manifest contains inputs that are not part of the current partition: "
            f"{sorted(unexpected_names)[:10]}"
        )
    if not dry_run:
        pd.DataFrame(manifest_rows).to_csv(dataset_manifest_path(dataset), index=False)
    return paths


def select_batch_paths(
    dataset: str, selected_batch: str | None, dry_run: bool = False
) -> list[Path]:
    paths = sync_inputs(dataset, dry_run=dry_run)
    if selected_batch:
        pattern = batch_selector_pattern(selected_batch)
        paths = [path for path in paths if pattern.fullmatch(path.name)]
        if not paths:
            raise FileNotFoundError(f"Critic batch not found for {dataset}: {selected_batch}")
    return paths


def canonical_dir(dataset: str, model: str) -> Path:
    return GEMINI_DIR / dataset / "outputs" / model_slug(model)


def canonical_path(dataset: str, model: str, input_path: Path) -> Path:
    return canonical_dir(dataset, model) / input_path.name


def run_dir(dataset: str, model: str, input_path: Path) -> Path:
    return GEMINI_DIR / dataset / "runs" / model_slug(model) / input_path.stem


def task_prompt(input_path: Path) -> str:
    codebook = CODEBOOK_PATH.read_text(encoding="utf-8")
    critic_prompt = PROMPT_PATH.read_text(encoding="utf-8")
    batch_content = input_path.read_text(encoding="utf-8")
    rows = validate_critic_input(input_path)
    output_template = {
        "criticisms": [
            {
                "annotation_id": "<preserve exactly>",
                "overall_error_prob": 0.0,
                "substantive_error_prob": 0.0,
                "clinical_error_prob": 0.0,
                "lived_error_prob": 0.0,
                "substantive_suggested_label": True,
                "clinical_suggested_label": True,
                "lived_suggested_label": False,
                "reason": "<one short sentence>",
                "evidence": "<shortest relevant excerpt>",
            }
        ]
    }
    return (
        f"{codebook}\n\n---\n\n"
        f"{critic_prompt}\n\n---\n\n"
        "Critique exactly one hierarchical ADHD/autism frame-annotation batch.\n\n"
        f"Input batch: `{input_path.relative_to(PROJECT_ROOT)}`\n"
        f"Expected rows: {len(rows)}\n"
        f"Input SHA-256: `{sha256(input_path)}`\n\n"
        f"Batch content:\n{batch_content}\n\n"
        "Return only one valid JSON object with exactly this shape and these field names. "
        "Return one criticism per input row in original order. Do not call tools, write files, "
        "or describe completion. Use null where the critic instructions require it. "
        "Keep reason at no more than 240 characters and evidence at no more than 180 characters. "
        "Use genuinely calibrated probabilities: reserve exactly 0 and 1 for logical certainty, "
        "and distinguish plausible boundary cases with non-extreme values.\n"
        + json.dumps(output_template, separators=(",", ":"))
    )


def gemini_command(model: str, prompt: str) -> list[str]:
    return [
        "gemini",
        "--prompt",
        prompt,
        "--model",
        model,
        "--output-format",
        "json",
        "--approval-mode",
        "default",
        "--policy",
        str(NO_TOOLS_POLICY_PATH),
    ]


def parse_json_object(text: str, source_name: str) -> dict[str, Any]:
    text = text.strip()
    fenced = re.fullmatch(r"```(?:json)?\s*(.*?)\s*```", text, flags=re.IGNORECASE | re.DOTALL)
    if fenced:
        text = fenced.group(1)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as error:
        raise ValueError(f"{source_name}: Gemini response is not valid JSON: {error}") from error
    if not isinstance(parsed, dict):
        raise ValueError(f"{source_name}: expected a JSON object.")
    return parsed


def parse_gemini_response(path: Path) -> tuple[Any, dict[str, Any]]:
    response = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(response, dict):
        raise ValueError(f"{path.name}: expected JSON object from Gemini CLI.")
    error = response.get("error")
    if error:
        raise NonRetryableGeminiError(f"Gemini CLI error: {error}")
    stats = response.get("stats", {})
    tools = stats.get("tools", {})
    tool_records = tools.get("byName", {})
    unexpected_successes = sorted(
        name
        for name, record in tool_records.items()
        if name != "update_topic" and record.get("success", 0)
    )
    if unexpected_successes:
        raise ValueError(
            f"{path.name}: Gemini successfully used excluded tools: {unexpected_successes}."
        )
    files = stats.get("files", {})
    if files.get("totalLinesAdded", 0) or files.get("totalLinesRemoved", 0):
        raise ValueError(f"{path.name}: Gemini modified files; expected no file changes.")
    text = response.get("response")
    if not isinstance(text, str) or not text.strip():
        raise ValueError(f"{path.name}: Gemini CLI returned no final response text.")
    structured = parse_json_object(text, path.name)
    if not isinstance(structured, dict) or set(structured) != {"criticisms"}:
        raise ValueError(f"{path.name}: model output does not match criticisms schema.")
    return structured["criticisms"], response


def read_gemini_response(path: Path) -> dict[str, Any] | None:
    if not path.exists() or path.stat().st_size == 0:
        return None
    try:
        response = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise ValueError(f"{path.name}: Gemini CLI response is empty or malformed JSON.") from error
    return response if isinstance(response, dict) else None


def gemini_version() -> str:
    result = subprocess.run(
        ["gemini", "--version"], cwd=PROJECT_ROOT, capture_output=True, text=True, check=False
    )
    return (result.stdout or result.stderr).strip()


def expected_contract(dataset: str, model: str, input_path: Path) -> dict[str, Any]:
    return {
        "dataset": dataset,
        "input_sha256": sha256(input_path),
        "codebook_version": CODEBOOK_VERSION,
        "codebook_sha256": sha256(CODEBOOK_PATH),
        "prompt_version": PROMPT_VERSION,
        "prompt_sha256": sha256(PROMPT_PATH),
        "schema_sha256": sha256(SCHEMA_PATH),
        "no_tools_policy_sha256": sha256(NO_TOOLS_POLICY_PATH),
        "task_prompt_sha256": hashlib.sha256(task_prompt(input_path).encode("utf-8")).hexdigest(),
        "requested_model": model,
    }


def refresh_run_manifest() -> None:
    rows = []
    for path in sorted(GEMINI_DIR.glob("*/runs/*/*/attempt_*_metadata.json")):
        metadata = json.loads(path.read_text(encoding="utf-8"))
        usage = metadata.pop("usage", {})
        metadata.pop("command", None)
        rows.append(
            {
                **metadata,
                **{f"usage_{key}": value for key, value in usage.items()},
                "metadata_path": str(path.relative_to(PROJECT_ROOT)),
            }
        )
    if rows:
        pd.DataFrame(rows).to_csv(GEMINI_DIR / "run_manifest.csv", index=False)


def write_metadata(
    path: Path,
    *,
    dataset: str,
    model: str,
    input_path: Path,
    command: list[str],
    attempt: int,
    return_code: int,
    status: str,
    response: dict[str, Any] | None = None,
    error: str = "",
    normalised_text_fields: int = 0,
    normalised_hierarchy_fields: int = 0,
) -> None:
    metadata = {
        **expected_contract(dataset, model, input_path),
        "batch_name": input_path.name,
        "input_rows": len(validate_critic_input(input_path)),
        "gemini_cli_version": gemini_version(),
        "command": command,
        "attempt": attempt,
        "return_code": return_code,
        "status": status,
        "error": error,
        "normalised_text_fields": normalised_text_fields,
        "normalised_hierarchy_fields": normalised_hierarchy_fields,
        "usage": response.get("stats", {}) if response else {},
        "completed_at_utc": utc_now(),
    }
    path.write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    refresh_run_manifest()


def validated_canonical_exists(dataset: str, model: str, input_path: Path) -> bool:
    canonical = canonical_path(dataset, model, input_path)
    if not canonical.exists():
        return False
    try:
        validate_critic_rows(
            read_jsonl(canonical), validate_critic_input(input_path), canonical.name
        )
    except ValueError:
        return False
    contract = expected_contract(dataset, model, input_path)
    metadata_paths = sorted(
        run_dir(dataset, model, input_path).glob("attempt_*_metadata.json")
    )
    for metadata_path in reversed(metadata_paths):
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if metadata.get("status") != "valid":
            continue
        return all(metadata.get(key) == value for key, value in contract.items())
    return False


def recover_valid_attempt(dataset: str, model: str, input_path: Path) -> int | None:
    expected_rows = validate_critic_input(input_path)
    contract = expected_contract(dataset, model, input_path)
    destination = run_dir(dataset, model, input_path)
    for response_path in sorted(destination.glob("attempt_*_response.json"), reverse=True):
        attempt = int(response_path.name.split("_")[1])
        metadata_path = destination / f"attempt_{attempt:02d}_metadata.json"
        if not metadata_path.exists():
            continue
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if not all(metadata.get(key) == value for key, value in contract.items()):
            continue
        try:
            criticisms, _ = parse_gemini_response(response_path)
            criticisms, normalised_text_fields = normalise_critic_text(criticisms)
            criticisms, normalised_hierarchy_fields = normalise_hierarchy_nulls(
                criticisms, expected_rows
            )
            validated = validate_critic_rows(criticisms, expected_rows, response_path.name)
        except (json.JSONDecodeError, OSError, RuntimeError, ValueError):
            continue
        write_jsonl(validated, canonical_path(dataset, model, input_path))
        metadata["status"] = "valid"
        metadata["error"] = ""
        metadata["normalised_text_fields"] = normalised_text_fields
        metadata["normalised_hierarchy_fields"] = normalised_hierarchy_fields
        metadata["recovered_at_utc"] = utc_now()
        metadata_path.write_text(json.dumps(metadata, indent=2), encoding="utf-8")
        refresh_run_manifest()
        return attempt
    return None


def run_gemini_process(
    command: list[str],
    *,
    stdout: Any,
    stderr: Any,
    timeout_seconds: int,
) -> subprocess.CompletedProcess[Any]:
    started = time.monotonic()
    process = subprocess.Popen(
        command,
        cwd=PROJECT_ROOT,
        stdout=stdout,
        stderr=stderr,
        text=True,
    )
    while True:
        try:
            return_code = process.wait(timeout=min(60, timeout_seconds))
            return subprocess.CompletedProcess(command, return_code)
        except subprocess.TimeoutExpired:
            elapsed = int(time.monotonic() - started)
            if elapsed >= timeout_seconds:
                process.kill()
                process.wait()
                return subprocess.CompletedProcess(command, 124)
            print(
                f"    Gemini still running after {elapsed // 60}m {elapsed % 60:02d}s...",
                flush=True,
            )


def run_batch(
    dataset: str,
    model: str,
    input_path: Path,
    max_attempts: int,
    resume: bool,
    timeout_seconds: int,
) -> str:
    if resume and validated_canonical_exists(dataset, model, input_path):
        return "skipped valid"
    if resume and (recovered_attempt := recover_valid_attempt(dataset, model, input_path)):
        return f"recovered valid attempt {recovered_attempt}"
    destination = run_dir(dataset, model, input_path)
    destination.mkdir(parents=True, exist_ok=True)
    existing_attempts = max(
        (
            int(path.name.split("_")[1])
            for path in destination.glob("attempt_*_*")
            if path.name.split("_")[1].isdigit()
        ),
        default=0,
    )
    last_error = ""
    for attempt in range(existing_attempts + 1, existing_attempts + max_attempts + 1):
        stem = f"attempt_{attempt:02d}"
        response_path = destination / f"{stem}_response.json"
        stderr_path = destination / f"{stem}_stderr.log"
        metadata_path = destination / f"{stem}_metadata.json"
        command = gemini_command(model, task_prompt(input_path))
        with (
            response_path.open("w", encoding="utf-8") as stdout,
            stderr_path.open("w", encoding="utf-8") as stderr,
        ):
            result = run_gemini_process(
                command,
                stdout=stdout,
                stderr=stderr,
                timeout_seconds=timeout_seconds,
            )
        response: dict[str, Any] | None = None
        try:
            response = read_gemini_response(response_path)
            if response and response.get("error"):
                raise NonRetryableGeminiError(f"Gemini CLI error: {response['error']}")
            if result.returncode == 124:
                raise NonRetryableGeminiError(
                    f"Gemini call exceeded the {timeout_seconds}-second timeout."
                )
            if result.returncode != 0:
                stderr_text = stderr_path.read_text(encoding="utf-8").strip()
                raise RuntimeError(
                    f"gemini exited with status {result.returncode}: {stderr_text[:300]}"
                )
            criticisms, response = parse_gemini_response(response_path)
            criticisms, normalised_text_fields = normalise_critic_text(criticisms)
            expected_rows = validate_critic_input(input_path)
            criticisms, normalised_hierarchy_fields = normalise_hierarchy_nulls(
                criticisms, expected_rows
            )
            validated = validate_critic_rows(
                criticisms, expected_rows, response_path.name
            )
            write_jsonl(validated, canonical_path(dataset, model, input_path))
            write_metadata(
                metadata_path,
                dataset=dataset,
                model=model,
                input_path=input_path,
                command=command,
                attempt=attempt,
                return_code=result.returncode,
                status="valid",
                response=response,
                normalised_text_fields=normalised_text_fields,
                normalised_hierarchy_fields=normalised_hierarchy_fields,
            )
            return f"valid attempt {attempt}"
        except (json.JSONDecodeError, OSError, RuntimeError, ValueError) as error:
            last_error = str(error)
            write_metadata(
                metadata_path,
                dataset=dataset,
                model=model,
                input_path=input_path,
                command=command,
                attempt=attempt,
                return_code=result.returncode if result else -1,
                status="invalid",
                response=response,
                error=last_error,
            )
            if isinstance(error, NonRetryableGeminiError):
                raise RuntimeError(
                    f"{input_path.name} stopped without retry: {last_error}"
                ) from error
    raise RuntimeError(f"{input_path.name} failed after {max_attempts} attempts: {last_error}")


def validate_dataset(dataset: str, model: str, selected_batch: str | None) -> None:
    missing = []
    for input_path in select_batch_paths(dataset, selected_batch):
        canonical = canonical_path(dataset, model, input_path)
        if not validated_canonical_exists(dataset, model, input_path):
            print(f"missing, invalid, or stale: {canonical.relative_to(PROJECT_ROOT)}")
            missing.append(canonical)
        else:
            print(f"valid: {canonical.relative_to(PROJECT_ROOT)}")
    if missing:
        raise FileNotFoundError(f"Missing {len(missing)} expected canonical critic output(s).")


def combine_dataset(dataset: str, model: str) -> Path:
    combined = []
    for input_path in select_batch_paths(dataset, None):
        canonical = canonical_path(dataset, model, input_path)
        if not validated_canonical_exists(dataset, model, input_path):
            raise ValueError(f"Missing, invalid, or stale canonical output: {canonical}")
        for row in read_jsonl(canonical):
            combined.append({**row, "source_file": input_path.name})
    output_dir = GEMINI_DIR / dataset / "scores" / model_slug(model)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"frame_llm_critic_scores_{dataset}.csv"
    pd.DataFrame(combined).to_csv(output_path, index=False)
    print(f"Wrote {len(combined)} critic scores: {output_path.relative_to(PROJECT_ROOT)}")
    return output_path


def pilot_truth() -> pd.DataFrame:
    human = pd.read_excel(PILOT_SOURCE)
    labels = pd.read_csv(
        CODEX_PILOT_DIR / "evaluation" / f"pilot_codex_labels_{ANNOTATOR_REASONING}.csv"
    )
    merged = human.merge(labels, on="annotation_id", suffixes=("_human", "_codex"))
    for axis in LABEL_COLUMNS:
        merged[f"{axis}_human_normalised"] = merged[f"{axis}_human"].map(parse_bool_or_none)
        merged[f"{axis}_codex_normalised"] = merged[f"{axis}_codex"].map(parse_bool_or_none)
    merged["any_hierarchical_error"] = [
        any(human_value != codex_value for human_value, codex_value in zip(human_row, codex_row))
        for human_row, codex_row in zip(
            merged[[f"{axis}_human_normalised" for axis in LABEL_COLUMNS]].itertuples(
                index=False, name=None
            ),
            merged[[f"{axis}_codex_normalised" for axis in LABEL_COLUMNS]].itertuples(
                index=False, name=None
            ),
        )
    ]
    return merged


def evaluate_pilot(models: list[str], top_k: int) -> None:
    truth = pilot_truth()[["annotation_id", "any_hierarchical_error"]]
    rows = []
    comparisons = []
    for model in models:
        scores_path = combine_dataset("pilot", model)
        scores = pd.read_csv(scores_path)
        comparison = truth.merge(scores, on="annotation_id", validate="one_to_one")
        comparison = comparison.sort_values(
            ["overall_error_prob", "annotation_id"], ascending=[False, True]
        ).reset_index(drop=True)
        positives = int(comparison["any_hierarchical_error"].sum())
        found = int(comparison.head(top_k)["any_hierarchical_error"].sum())
        recall = found / positives if positives else 1.0
        ap = average_precision_score(
            comparison["any_hierarchical_error"].astype(int), comparison["overall_error_prob"]
        )
        rows.append(
            {
                "requested_model": model,
                "rows": len(comparison),
                "known_errors": positives,
                f"errors_found_top_{top_k}": found,
                f"recall_top_{top_k}": recall,
                "average_precision": ap,
                "passes_top_k_recall_gate": recall >= 0.60,
            }
        )
        comparison["model"] = model
        comparisons.append(comparison)
    GEMINI_PILOT_EVAL.mkdir(parents=True, exist_ok=True)
    metrics = pd.DataFrame(rows).sort_values(
        [f"errors_found_top_{top_k}", "average_precision"], ascending=False
    )
    metrics.to_csv(GEMINI_PILOT_EVAL / "critic_pilot_model_metrics.csv", index=False)
    pd.concat(comparisons, ignore_index=True).to_csv(
        GEMINI_PILOT_EVAL / "critic_pilot_comparisons.csv", index=False
    )
    print(metrics.to_string(index=False))


def rank_production_scores(model: str) -> pd.DataFrame:
    scores = pd.read_csv(combine_dataset("production", model))
    axis_columns = ["substantive_error_prob", "clinical_error_prob", "lived_error_prob"]
    scores["max_applicable_axis_error_prob"] = scores[axis_columns].max(axis=1, skipna=True)
    return scores.sort_values(
        ["overall_error_prob", "max_applicable_axis_error_prob", "annotation_id"],
        ascending=[False, False, True],
    ).reset_index(drop=True)


def review_frame(model: str) -> pd.DataFrame:
    scores = rank_production_scores(model)
    labels = pd.read_csv(ANNOTATOR_LABELS)
    pool = pd.read_csv(ANNOTATOR_POOL)
    scores["rank"] = range(1, len(scores) + 1)
    frame = labels.merge(scores, on="annotation_id", validate="one_to_one")
    if len(frame) != len(scores) or len(frame) != len(labels):
        raise ValueError(
            "Combined Codex labels, critic scores, and production pool must cover identical IDs."
        )
    frame = frame.merge(
        pool[
            [
                "annotation_id",
                "context_id",
                "analysis_unit",
                "lsc_year",
                "raw_form",
                "target_sentence_plus_adjacent",
            ]
        ],
        on="annotation_id",
        validate="one_to_one",
    )
    frame = frame.sort_values("rank").reset_index(drop=True)
    frame["final_substantive_target_discourse"] = (
        frame["substantive_target_discourse"].map(parse_bool_or_none).map(label_text)
    )
    frame["final_clinical_frame_present"] = (
        frame["clinical_frame_present"].map(parse_bool_or_none).map(label_text)
    )
    frame["final_lived_experience_frame_present"] = (
        frame["lived_experience_frame_present"].map(parse_bool_or_none).map(label_text)
    )
    frame["review_status"] = "pending"
    frame["adjudication_note"] = ""
    return frame


def write_review_workbook(frame: pd.DataFrame, path: Path, sheet_name: str) -> None:
    columns = [
        "rank",
        "annotation_id",
        "context_id",
        "analysis_unit",
        "lsc_year",
        "raw_form",
        "target_sentence_plus_adjacent",
        "substantive_target_discourse",
        "clinical_frame_present",
        "lived_experience_frame_present",
        "confidence",
        "overall_error_prob",
        "substantive_error_prob",
        "clinical_error_prob",
        "lived_error_prob",
        "substantive_suggested_label",
        "clinical_suggested_label",
        "lived_suggested_label",
        "reason",
        "evidence",
        "final_substantive_target_discourse",
        "final_clinical_frame_present",
        "final_lived_experience_frame_present",
        "review_status",
        "adjudication_note",
    ]
    if "review_wave" in frame:
        columns.insert(1, "review_wave")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    header_fill = PatternFill("solid", fgColor="1F4E78")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    for column_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.protection = Protection(locked=True)
    for row_index, row in enumerate(frame[columns].itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            column_name = columns[column_index - 1]
            if column_name.startswith("final_") or column_name in {
                "review_status",
                "adjudication_note",
            }:
                cell.protection = Protection(locked=False)
                cell.fill = editable_fill
            else:
                cell.protection = Protection(locked=True)
    label_validation = DataValidation(type="list", formula1='"TRUE,FALSE,NA"', allow_blank=False)
    status_validation = DataValidation(
        type="list", formula1='"pending,reviewed"', allow_blank=False
    )
    sheet.add_data_validation(label_validation)
    sheet.add_data_validation(status_validation)
    if not frame.empty:
        for column in [
            "final_substantive_target_discourse",
            "final_clinical_frame_present",
            "final_lived_experience_frame_present",
        ]:
            index = columns.index(column) + 1
            label_validation.add(
                f"{get_column_letter(index)}2:{get_column_letter(index)}{len(frame) + 1}"
            )
        status_index = columns.index("review_status") + 1
        status_validation.add(
            f"{get_column_letter(status_index)}2:{get_column_letter(status_index)}{len(frame) + 1}"
        )
    widths = {
        "annotation_id": 20,
        "context_id": 24,
        "target_sentence_plus_adjacent": 80,
        "reason": 40,
        "evidence": 40,
        "adjudication_note": 40,
    }
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(column, 18)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.protection.sheet = True
    sheet.protection.password = "frame"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def prepare_ranked_review(model: str, ranked_limit: int, wave_size: int) -> Path:
    frame = review_frame(model).head(ranked_limit).copy()
    frame["review_wave"] = ((frame["rank"] - 1) // wave_size) + 1
    path = CORRECTION_DIR / "frame_llm_ranked_review.xlsx"
    if path.exists():
        raise FileExistsError(f"Refusing to overwrite an existing human-review workbook: {path}")
    write_review_workbook(frame, path, "ranked_review")
    print(f"Wrote {len(frame)} ranked review rows: {path.relative_to(PROJECT_ROOT)}")
    return path


def read_review_workbook(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Review workbook not found: {path}")
    frame = pd.read_excel(path)
    if "review_status" not in frame:
        raise ValueError(f"{path.name}: missing review_status.")
    invalid = set(frame["review_status"].dropna().astype(str)) - REVIEW_STATUS_VALUES
    if invalid:
        raise ValueError(f"{path.name}: invalid review statuses: {sorted(invalid)}")
    return frame


def row_has_correction(row: pd.Series) -> bool:
    return any(
        parse_bool_or_none(row[f"final_{axis}"]) != parse_bool_or_none(row[axis])
        for axis in LABEL_COLUMNS
    )


def validate_ranked_review_completion(ranked: pd.DataFrame) -> set[str]:
    ranked = ranked.sort_values("rank").reset_index(drop=True)
    reviewed_mask = ranked["review_status"].eq("reviewed")
    reviewed_rows = int(reviewed_mask.sum())
    if reviewed_rows == 0:
        raise ValueError("No ranked-review rows are marked reviewed.")
    if reviewed_mask.tolist() != [True] * reviewed_rows + [False] * (len(ranked) - reviewed_rows):
        raise ValueError("Ranked review must be a contiguous prefix in score order.")
    if reviewed_rows < len(ranked):
        if reviewed_rows < 200 or reviewed_rows % 100:
            raise ValueError(
                "Early ranked-review stopping is permitted only after complete 100-case waves."
            )
        reviewed = ranked.head(reviewed_rows).copy()
        reviewed["any_correction"] = reviewed.apply(row_has_correction, axis=1)
        previous_yield = int(reviewed.iloc[-200:-100]["any_correction"].sum())
        latest_yield = int(reviewed.iloc[-100:]["any_correction"].sum())
        if previous_yield >= 5 or latest_yield >= 5:
            raise ValueError(
                "Early stopping requires fewer than five corrections in each of the last "
                "two complete 100-case waves."
            )
    return set(ranked.loc[reviewed_mask, "annotation_id"].astype(str))


def prepare_audit(model: str, rows: int, seed: int) -> Path:
    ranked_path = CORRECTION_DIR / "frame_llm_ranked_review.xlsx"
    ranked = read_review_workbook(ranked_path)
    reviewed_ids = validate_ranked_review_completion(ranked)
    candidates = review_frame(model)
    candidates = candidates.loc[~candidates["annotation_id"].astype(str).isin(reviewed_ids)]
    if rows > len(candidates):
        raise ValueError(f"Requested {rows} audit rows but only {len(candidates)} remain.")
    audit = candidates.sample(n=rows, random_state=seed).sort_values("rank")
    path = CORRECTION_DIR / "frame_llm_residual_audit.xlsx"
    if path.exists():
        raise FileExistsError(f"Refusing to overwrite an existing human-review workbook: {path}")
    write_review_workbook(audit, path, "residual_audit")
    print(f"Wrote {len(audit)} residual audit rows: {path.relative_to(PROJECT_ROOT)}")
    return path


def corrected_labels(row: pd.Series) -> tuple[bool, bool | None, bool | None]:
    labels = (
        parse_bool_or_none(row["final_substantive_target_discourse"]),
        parse_bool_or_none(row["final_clinical_frame_present"]),
        parse_bool_or_none(row["final_lived_experience_frame_present"]),
    )
    if labels[0] is None:
        raise ValueError(f"{row['annotation_id']}: final substantive label cannot be NA.")
    validate_hierarchy(*labels)
    return labels


def finalize_corrections() -> Path:
    labels = pd.read_csv(ANNOTATOR_LABELS)
    pool = pd.read_csv(ANNOTATOR_POOL)
    ranked = read_review_workbook(CORRECTION_DIR / "frame_llm_ranked_review.xlsx")
    validate_ranked_review_completion(ranked)
    audit = read_review_workbook(CORRECTION_DIR / "frame_llm_residual_audit.xlsx")
    if audit.empty or not audit["review_status"].eq("reviewed").all():
        raise ValueError("Every residual-audit row must be reviewed before finalization.")
    reviews = [
        ranked.loc[ranked["review_status"].eq("reviewed")].copy(),
        audit.copy(),
    ]
    reviewed = pd.concat(reviews, ignore_index=True) if reviews else pd.DataFrame()
    if not reviewed.empty and reviewed["annotation_id"].duplicated().any():
        raise ValueError("An annotation appears as reviewed in both review workbooks.")
    unknown_review_ids = set(reviewed["annotation_id"].astype(str)) - set(
        labels["annotation_id"].astype(str)
    )
    if unknown_review_ids:
        raise ValueError(
            f"Review workbooks contain unknown annotation IDs: {sorted(unknown_review_ids)[:10]}"
        )
    reviewed_by_id = reviewed.set_index("annotation_id") if not reviewed.empty else pd.DataFrame()

    output = labels.merge(
        pool[
            [
                "annotation_id",
                "context_id",
                "analysis_unit",
                "lsc_year",
                "raw_form",
                "target_sentence_plus_adjacent",
            ]
        ],
        on="annotation_id",
        validate="one_to_one",
    )
    corrected_rows = []
    for row in output.itertuples(index=False):
        annotation_id = str(row.annotation_id)
        if not reviewed.empty and annotation_id in reviewed_by_id.index:
            final = corrected_labels(reviewed_by_id.loc[annotation_id])
            correction_source = "human_review"
        else:
            final = (
                parse_bool_or_none(row.substantive_target_discourse),
                parse_bool_or_none(row.clinical_frame_present),
                parse_bool_or_none(row.lived_experience_frame_present),
            )
            validate_hierarchy(*final)
            correction_source = "codex_unreviewed"
        corrected_rows.append(
            {
                "annotation_id": annotation_id,
                "context_id": row.context_id,
                "analysis_unit": row.analysis_unit,
                "lsc_year": row.lsc_year,
                "raw_form": row.raw_form,
                "target_sentence_plus_adjacent": row.target_sentence_plus_adjacent,
                "corrected_substantive_target_discourse": final[0],
                "corrected_clinical_frame_present": final[1],
                "corrected_lived_experience_frame_present": final[2],
                "corrected_confidence": row.confidence,
                "corrected_derived_frame": derive_frame(*final),
                "correction_source": correction_source,
                "codebook_version": CODEBOOK_VERSION,
                "prompt_version": f"annotator_v4+{PROMPT_VERSION}",
            }
        )
    result = pd.DataFrame(corrected_rows)
    if len(result) != len(labels) or result["annotation_id"].duplicated().any():
        raise ValueError("Final corrected labels do not match the complete annotation set.")
    path = CORRECTION_DIR / "frame_llm_correction_completed.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    result.to_csv(path, index=False)
    print(f"Wrote {len(result)} corrected training labels: {path.relative_to(PROJECT_ROOT)}")
    return path


def rebuild_batch(dataset: str, selected_batch: str) -> None:
    pattern = batch_selector_pattern(selected_batch)
    input_paths = [
        path
        for path in dataset_input_dir(dataset).glob("critic_batch_*.jsonl")
        if pattern.fullmatch(path.name)
    ]
    if not input_paths:
        raise FileNotFoundError(f"No critic inputs match: {selected_batch}")
    names = {path.name for path in input_paths}
    stems = {path.stem for path in input_paths}
    for input_path in input_paths:
        input_path.unlink()
    for root in [
        GEMINI_DIR / dataset / "outputs",
        GEMINI_DIR / dataset / "runs",
    ]:
        for path in (
            [
                candidate
                for candidate in root.glob("**/*")
                if candidate.name in names or candidate.name in stems
            ]
            if root.exists()
            else []
        ):
            shutil.rmtree(path) if path.is_dir() else path.unlink()
    manifest_path = dataset_manifest_path(dataset)
    manifest = pd.read_csv(manifest_path)
    manifest = manifest.loc[~manifest["critic_batch_name"].isin(names)]
    manifest.to_csv(manifest_path, index=False)
    sync_inputs(dataset)
    print(f"Deleted stale artifacts and rebuilt {len(names)} critic chunk(s).")


def probe(model: str) -> None:
    """Run a minimal call to inspect the raw Gemini CLI JSON envelope format."""
    cmd = gemini_command(
        model,
        'Return exactly this JSON object and nothing else: {"test": true}',
    )
    print("Command:", " ".join(cmd))
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60, cwd=PROJECT_ROOT)
    print("=== STDOUT ===")
    print(result.stdout)
    print("=== STDERR ===")
    print(result.stderr)
    print("=== Return code:", result.returncode)


def self_test() -> None:
    proposed_true = {
        "annotation_id": "test_1",
        "context_id": "ctx",
        "target": "ADHD",
        "raw_form": "ADHD",
        "passage": "A passage.",
        "proposed_substantive_target_discourse": True,
        "proposed_clinical_frame_present": True,
        "proposed_lived_experience_frame_present": False,
        "proposed_confidence": "high",
    }
    valid = {
        "annotation_id": "test_1",
        "overall_error_prob": 0.1,
        "substantive_error_prob": 0.1,
        "clinical_error_prob": 0.1,
        "lived_error_prob": 0.1,
        "substantive_suggested_label": True,
        "clinical_suggested_label": True,
        "lived_suggested_label": False,
        "reason": "Consistent.",
        "evidence": "A passage.",
    }
    validate_critic_rows([valid], [proposed_true], "valid")
    proposed_false = {
        **proposed_true,
        "annotation_id": "test_2",
        "proposed_substantive_target_discourse": False,
        "proposed_clinical_frame_present": None,
        "proposed_lived_experience_frame_present": None,
    }
    valid_false = {
        **valid,
        "annotation_id": "test_2",
        "clinical_error_prob": None,
        "lived_error_prob": None,
        "substantive_suggested_label": False,
        "clinical_suggested_label": None,
        "lived_suggested_label": None,
    }
    validate_critic_rows([valid_false], [proposed_false], "valid_false")
    invalid_cases = [
        {**valid, "clinical_error_prob": None},
        {**valid, "substantive_suggested_label": False},
        {**valid, "overall_error_prob": 1.1},
        {**valid, "annotation_id": "wrong"},
    ]
    for index, invalid in enumerate(invalid_cases, start=1):
        try:
            validate_critic_rows([invalid], [proposed_true], f"invalid_{index}")
        except ValueError:
            continue
        raise AssertionError(f"Validator accepted invalid self-test case {index}.")
    try:
        validate_critic_rows(
            [{**valid_false, "clinical_error_prob": 0.1}],
            [proposed_false],
            "invalid_non_substantive_score",
        )
    except ValueError:
        pass
    else:
        raise AssertionError("Validator accepted a Stage-1 score for a non-substantive proposal.")
    clipped, clipped_fields = normalise_critic_text([{**valid, "evidence": "x" * 181}])
    if clipped_fields != 1 or len(clipped[0]["evidence"]) != 180:
        raise AssertionError("Text normalisation did not clip evidence deterministically.")
    masked, masked_fields = normalise_hierarchy_nulls(
        [{**valid_false, "clinical_error_prob": 0.2, "lived_error_prob": 0.2}],
        [proposed_false],
    )
    if masked_fields != 2 or masked[0]["clinical_error_prob"] is not None:
        raise AssertionError("Hierarchy normalisation did not mask inapplicable probabilities.")
    if parse_json_object('```json\n{"criticisms":[]}\n```', "self_test") != {"criticisms": []}:
        raise AssertionError("Fenced JSON parser failed.")
    if parse_bool_or_none(1.0) is not True or parse_bool_or_none(0.0) is not False:
        raise AssertionError("Numeric workbook labels were not parsed as booleans.")
    try:
        parse_bool_or_none(0.5)
    except ValueError:
        pass
    else:
        raise AssertionError("Parser accepted an invalid numeric label.")

    review_rows = pd.DataFrame(
        [
            {
                "rank": rank,
                "annotation_id": f"review_{rank}",
                "review_status": "reviewed" if rank <= 200 else "pending",
                "substantive_target_discourse": True,
                "clinical_frame_present": True,
                "lived_experience_frame_present": False,
                "final_substantive_target_discourse": "TRUE",
                "final_clinical_frame_present": "TRUE",
                "final_lived_experience_frame_present": "FALSE",
            }
            for rank in range(1, 301)
        ]
    )
    validate_ranked_review_completion(review_rows)
    review_rows.loc[100, "review_status"] = "pending"
    try:
        validate_ranked_review_completion(review_rows)
    except ValueError:
        pass
    else:
        raise AssertionError("Review validator accepted a gap in the ranked-review prefix.")
    print(f"Validator self-test passed: 3 valid and {len(invalid_cases) + 2} invalid cases.")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run reproducible, schema-validated Gemini frame criticism."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)
    sync = subparsers.add_parser("sync")
    sync.add_argument("--dataset", choices=["pilot", "production"], required=True)
    for name in ["dry-run", "run", "validate"]:
        command = subparsers.add_parser(name)
        command.add_argument("--dataset", choices=["pilot", "production"], required=True)
        command.add_argument("--model", default=DEFAULT_MODEL)
        command.add_argument("--batch")
        if name == "run":
            command.add_argument("--max-attempts", type=int, default=1)
            command.add_argument("--no-resume", action="store_true")
            command.add_argument("--timeout-seconds", type=int, default=1200)
    combine = subparsers.add_parser("combine")
    combine.add_argument("--dataset", choices=["pilot", "production"], required=True)
    combine.add_argument("--model", default=DEFAULT_MODEL)
    evaluate = subparsers.add_parser("evaluate-pilot")
    evaluate.add_argument("--model", action="append", required=True)
    evaluate.add_argument("--top-k", type=int, default=40)
    review = subparsers.add_parser("prepare-ranked-review")
    review.add_argument("--model", default=DEFAULT_MODEL)
    review.add_argument("--ranked-limit", type=int, default=550)
    review.add_argument("--wave-size", type=int, default=100)
    audit = subparsers.add_parser("prepare-audit")
    audit.add_argument("--model", default=DEFAULT_MODEL)
    audit.add_argument("--rows", type=int, default=50)
    audit.add_argument("--seed", type=int, default=20260611)
    subparsers.add_parser("finalize-corrections")
    rebuild = subparsers.add_parser("rebuild-batch")
    rebuild.add_argument("--dataset", choices=["pilot", "production"], required=True)
    rebuild.add_argument("--batch", required=True)
    probe_cmd = subparsers.add_parser("probe")
    probe_cmd.add_argument("--model", default=DEFAULT_MODEL)
    subparsers.add_parser("self-test")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.command == "self-test":
        self_test()
    elif args.command == "sync":
        paths = sync_inputs(args.dataset)
        print(f"Synchronized {len(paths)} {args.dataset} critic input batch(es).")
    elif args.command == "rebuild-batch":
        rebuild_batch(args.dataset, args.batch)
    elif args.command == "evaluate-pilot":
        evaluate_pilot(args.model, args.top_k)
    elif args.command == "prepare-ranked-review":
        prepare_ranked_review(args.model, args.ranked_limit, args.wave_size)
    elif args.command == "prepare-audit":
        prepare_audit(args.model, args.rows, args.seed)
    elif args.command == "finalize-corrections":
        finalize_corrections()
    elif args.command == "combine":
        combine_dataset(args.dataset, args.model)
    elif args.command == "validate":
        validate_dataset(args.dataset, args.model, args.batch)
    elif args.command == "probe":
        probe(args.model)
    elif args.command == "dry-run":
        paths = select_batch_paths(args.dataset, args.batch, dry_run=True)
        for path in paths:
            if not path.exists():
                print(f"would append critic input: {path.relative_to(PROJECT_ROOT)}")
            print(" ".join(gemini_command(args.model, "<prompt omitted>")))
    elif args.command == "run":
        paths = select_batch_paths(args.dataset, args.batch)
        for index, path in enumerate(paths, start=1):
            print(f"[{index}/{len(paths)}] {path.name}", flush=True)
            print(
                " ",
                run_batch(
                    args.dataset,
                    args.model,
                    path,
                    args.max_attempts,
                    not args.no_resume,
                    args.timeout_seconds,
                ),
            )


if __name__ == "__main__":
    main()
